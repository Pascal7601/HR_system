from decimal import Decimal
from app.models import Employee, Payslip, Leave, LeaveRequest
from app.extensions import db

from decimal import Decimal, ROUND_HALF_UP
from datetime import date, timedelta
import calendar

# ---------------------------------------------------------------------------
# ASSUMPTIONS (documented, not meant to match any real country's tax code):

# - Working days = Mon-Fri in the calendar month. Holidays are not excluded.
# - An employee is only paid from their hire_date onward if they joined
#   mid-period; weekdays before that date are not counted as "days owed."
# - Only leave of type "unpaid" reduces pay. Approved leave of any other
#   type (annual, sick) does NOT reduce paid_days.
# - Tax is calculated MARGINALLY: each bracket's rate applies only to the
#   slice of income within it, not the whole salary at one flat rate.
#
#   Monthly tax brackets (illustrative, not a real jurisdiction's rates):
#     0    - 30000   -> 0%
#     30000 - 100000   -> 10%
#     100000+         -> 20%
#
# - Social security = flat 5% of gross pay, capped at 500.
# - net_pay = gross_pay - tax - social_security
# ---------------------------------------------------------------------------

TAX_BRACKETS = [
    (Decimal("0"), Decimal("30000"), Decimal("0.00")),
    (Decimal("30000"), Decimal("100000"), Decimal("0.10")),
    (Decimal("100000"), None, Decimal("0.20")),  # None = no upper bound
]

SOCIAL_SECURITY_RATE = Decimal("0.05")
SOCIAL_SECURITY_CAP = Decimal("500")


def _round_money(value):
    return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

def apply_marginal_brackets(taxable_income):
    """Marginal tax calculation: each bracket's rate applies only to the
    slice of income that falls within that bracket, not the whole amount."""
    if taxable_income <= 0:
        return Decimal("0.00")

    tax = Decimal("0")
    for lower, upper, rate in TAX_BRACKETS:
        if taxable_income <= lower:
            break
        slice_top = taxable_income if upper is None else min(taxable_income, upper)
        taxable_slice = slice_top - lower
        if taxable_slice > 0:
            tax += taxable_slice * rate
    return _round_money(tax)

def _count_weekdays(start_date, end_date):
    """Count Mon-Fri days in [start_date, end_date], inclusive. Returns 0
    if start_date is after end_date (e.g. employee hired after this period)."""
    if start_date > end_date:
        return 0
    return sum(
        1
        for offset in range((end_date - start_date).days + 1)
        if (start_date + timedelta(days=offset)).weekday() < 5
    )

def _unpaid_leave_weekdays(employee_id, period_start, period_end):
    """Sum weekdays covered by approved 'unpaid' leave requests that overlap
    this period, clipped to the period's bounds."""
    unpaid_type = Leave.query.filter_by(name="unpaid").first()
    if not unpaid_type:
        return 0

    overlapping = LeaveRequest.query.filter(
        LeaveRequest.employee_id == employee_id,
        LeaveRequest.leave_type_id == unpaid_type.id,
        LeaveRequest.status == "approved",
        LeaveRequest.start_date <= period_end,
        LeaveRequest.end_date >= period_start,
    ).all()

    total = 0
    for req in overlapping:
        clipped_start = max(req.start_date, period_start)
        clipped_end = min(req.end_date, period_end)
        total += _count_weekdays(clipped_start, clipped_end)
    return total


def generate_payslip(data):
    """
    Generates a payslip for the specified employee and period.
    Args:
        data (dict): A dictionary containing the employee ID, period month, and period year.
    Returns:
        Payslip: The generated payslip object.

    """
    employee = Employee.query.get(str(data["employee_id"]))
    if not employee:
        return None, "Employee not found"

    month = data["period_month"]
    year = data["period_year"]

    existing = Payslip.query.filter_by(
        employee_id=employee.id, period_month=month, period_year=year
    ).first()
    if existing:
        return None, "Payslip already generated for this period"

    period_start = date(year, month, 1)
    period_end = date(year, month, calendar.monthrange(year, month)[1])

    working_days_in_period = _count_weekdays(period_start, period_end)
    if working_days_in_period == 0:
        return None, "Selected period has no working days"

    # Mid-month joiner: only count from hire_date onward, if it falls in this period
    effective_start = period_start
    if employee.hire_date and employee.hire_date > period_start:
        effective_start = employee.hire_date

    if effective_start > period_end:
        return None, "Employee's hire date is after this period; nothing to pay"

    days_present = _count_weekdays(effective_start, period_end)
    unpaid_days = _unpaid_leave_weekdays(employee.id, effective_start, period_end)

    paid_days = max(days_present - unpaid_days, 0)

    basic_salary = employee.salary or Decimal("0")
    gross_pay = _round_money(
        basic_salary * Decimal(paid_days) / Decimal(working_days_in_period)
    )

    tax_amount = apply_marginal_brackets(gross_pay)
    social_security_amount = _round_money(
        min(gross_pay * SOCIAL_SECURITY_RATE, SOCIAL_SECURITY_CAP)
    )
    net_pay = gross_pay - tax_amount - social_security_amount

    payslip = Payslip(
        employee_id=employee.id,
        period_month=month,
        period_year=year,
        working_days_in_period=working_days_in_period,
        paid_days=paid_days,
        gross_pay=gross_pay,
        tax_amount=tax_amount,
        social_security_amount=social_security_amount,
        net_pay=net_pay,
        status="finalized",
    )
    db.session.add(payslip)
    db.session.commit()
    return payslip
    

def get_payslips_for_employee(employee_id, month=None, year=None):
    """
    Retrieves all payslips for the specified employee.
    Args:
        employee_id (str): The ID of the employee.
    Returns:
        list: A list of payslip objects for the employee.
    """
    employee = Employee.query.filter_by(user_id=employee_id).first()
    if not employee:
        return []
    query = Payslip.query.filter_by(employee_id=employee.id)
    if month:
        query = query.filter_by(period_month=month)
    if year:
        query = query.filter_by(period_year=year)
    return query.order_by(Payslip.period_year.desc(), Payslip.period_month.desc()).all()

def generate_payslips_for_period(month, year):
    """Generate a payslip for every active employee for this period.
    Skips employees who already have one (idempotent) rather than erroring out."""

    employees = Employee.query.filter_by(employment_status="active").all()
    created, skipped = [], []

    for employee in employees:
        existing = Payslip.query.filter_by(
            employee_id=employee.id, period_month=month, period_year=year
        ).first()
        if existing:
            skipped.append(employee.staff_no)
            continue

        payslip = generate_payslip({
            "employee_id": employee.id,
            "period_month": month,
            "period_year": year,
        })
        if payslip:
            created.append(employee.staff_no)

    return {"created": created, "skipped": skipped}

def get_payslips_for_period(month, year):
    return Payslip.query.filter_by(period_month=month, period_year=year).order_by(
        Payslip.generated_at.desc()
    ).all()

import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def build_payslips_excel(month, year):
    """Build an in-memory .xlsx workbook of all payslips for a given period."""
    payslips = get_payslips_for_period(month, year)

    wb = Workbook()
    sheet = wb.active
    month_name = [
        "January", "February", "March", "April", "May", "June",
        "July", "August", "September", "October", "November", "December",
    ][month - 1]
    sheet.title = f"Payroll {month_name} {year}"[:31]  # Excel sheet name limit

    headers = [
        "Staff No", "Employee Name", "Working Days", "Paid Days",
        "Gross Pay", "Tax", "Social Security", "Net Pay", "Status",
    ]

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3D5A4C", end_color="3D5A4C", fill_type="solid")
    body_font = Font(name="Arial")
    currency_format = "$#,##0.00"

    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, payslip in enumerate(payslips, start=2):
        employee = payslip.employee
        row_values = [
            employee.staff_no if employee else "—",
            employee.first_name + " " + employee.last_name if employee else "Unknown",
            payslip.working_days_in_period,
            float(payslip.paid_days),
            float(payslip.gross_pay),
            float(payslip.tax_amount),
            float(payslip.social_security_amount),
            float(payslip.net_pay),
            payslip.status,
        ]
        for col_idx, value in enumerate(row_values, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font
            if col_idx in (5, 6, 7, 8):  # Gross Pay, Tax, Social Security, Net Pay
                cell.number_format = currency_format

    # Totals row
    if payslips:
        total_row = len(payslips) + 2
        sheet.cell(row=total_row, column=2, value="Total").font = Font(name="Arial", bold=True)
        for col_idx in (5, 6, 7, 8):
            col_letter = get_column_letter(col_idx)
            cell = sheet.cell(
                row=total_row, column=col_idx,
                value=f"=SUM({col_letter}2:{col_letter}{total_row - 1})",
            )
            cell.font = Font(name="Arial", bold=True)
            cell.number_format = currency_format

    # Reasonable column widths instead of Excel's cramped default
    widths = [16, 22, 13, 11, 14, 12, 15, 14, 12]
    for col_idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(col_idx)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer